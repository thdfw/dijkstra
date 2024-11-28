import matplotlib
import matplotlib.pyplot as plt
from matplotlib.colors import Normalize, ListedColormap, BoundaryNorm
import os
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.drawing.image import Image
from config import DParams
from dtypes import DNode, DEdge, DForecast

class DGraph():

    def __init__(self):
        self.params = DParams()
        self.forecasts = DForecast(self.params)
        print("Creating graph...")
        self.create_nodes()
        self.create_edges()
        print("Solving Dijkstra...")
        self.solve_dijkstra()
        self.plot()
        print("Exporting to Excel...")
        self.export_excel()

    def create_nodes(self):
        self.initial_node = DNode(0, self.params.initial_top_temp, self.params.initial_thermocline, self.params)
        self.nodes = {}
        for time_slice in range(self.params.horizon+1):
            self.nodes[time_slice] = [self.initial_node] if time_slice==0 else []
            self.nodes[time_slice].extend(
                DNode(time_slice, top_temp, thermocline, self.params)
                for top_temp in self.params.available_top_temps[1:]
                for thermocline in range(1,self.params.num_layers+1)
                if (time_slice, top_temp, thermocline) != (0, self.params.initial_top_temp, self.params.initial_thermocline)
            )
            # TODO: check if you need index
            self.nodes_by_energy = sorted(self.nodes[time_slice], key=lambda x: (x.energy, x.top_temp), reverse=True)
            for n in self.nodes[time_slice]:
                n.index = self.nodes_by_energy.index(n)+1

    def create_edges(self):
        self.edges = {}
        self.bottom_node = DNode(0,self.params.available_top_temps[0], 0, self.params)
        self.top_node = DNode(0, self.params.available_top_temps[-1], self.params.num_layers, self.params)
        self.energy_between_consecutive_states = (DNode(0,self.params.available_top_temps[0],2,self.params).energy 
                                                  - DNode(0,self.params.available_top_temps[0],1,self.params).energy)

        for h in range(self.params.horizon):
            
            for node_now in self.nodes[h]:
                self.edges[node_now] = []
                
                for node_next in self.nodes[h+1]:

                    # The energy difference between two states might be lower than energy between two nodes
                    losses = self.params.storage_losses_percent/100 * (node_now.energy-self.bottom_node.energy)
                    if self.forecasts.load[h]==0 and losses>0 and losses<self.energy_between_consecutive_states:
                        losses = self.energy_between_consecutive_states + 1/1e9

                    store_heat_in = node_next.energy - node_now.energy
                    hp_heat_out = store_heat_in + self.forecasts.load[h] + losses
                    
                    # This condition reduces the amount of times we need to compute the COP
                    if (hp_heat_out/self.params.max_cop <= self.params.max_hp_elec_in and
                        hp_heat_out/self.params.min_cop >= self.params.min_hp_elec_in):
                    
                        cop = self.params.COP(oat=self.forecasts.oat[h], lwt=node_next.top_temp)

                        if (hp_heat_out/cop <= self.params.max_hp_elec_in and 
                            hp_heat_out/cop >= self.params.min_hp_elec_in):

                            cost = self.forecasts.elec_price[h]/100 * hp_heat_out/cop

                            # If some of the load is satisfied by the storage
                            # Then it must satisfy the SWT requirement
                            if store_heat_in < 0:
                                if ((hp_heat_out < self.forecasts.load[h] and 
                                     self.forecasts.load[h] > 0)
                                     and
                                    (node_now.top_temp < self.forecasts.rswt[h] or 
                                     node_next.top_temp < self.forecasts.rswt[h])):
                                    # TODO: add a soft constraint (e.g. cost+=1e5)
                                    continue
                            
                            self.edges[node_now].append(DEdge(node_now, node_next, cost, hp_heat_out))

    def solve_dijkstra(self):
        for time_slice in range(self.params.horizon-1, -1, -1):
            for node in self.nodes[time_slice]:
                best_edge = min(self.edges[node], key=lambda e: e.head.pathcost + e.cost)
                if best_edge.hp_heat_out < 0: 
                    best_edge_neg = max([e for e in self.edges[node] if e.hp_heat_out<0], key=lambda e: e.hp_heat_out)
                    best_edge_pos = min([e for e in self.edges[node] if e.hp_heat_out>=0], key=lambda e: e.hp_heat_out)
                    best_edge = best_edge_pos if (-best_edge_neg.hp_heat_out >= best_edge_pos.hp_heat_out) else best_edge_neg
                node.pathcost = best_edge.head.pathcost + best_edge.cost
                node.next_node = best_edge.head
        # for a in self.edges[self.initial_node]:
        #     print(a)

    def plot(self):
        # Walk along the shortest path (sp)
        sp_top_temp = []
        sp_thermocline = []
        sp_hp_heat_out = []
        sp_stored_energy = []
        node_i = self.initial_node
        the_end = False
        while not the_end:
            if node_i.next_node is None:
                the_end = True
                sp_hp_heat_out.append(edge_i.hp_heat_out)
            else:
                edge_i = [e for e in self.edges[node_i] if e.head==node_i.next_node][0]
                sp_hp_heat_out.append(edge_i.hp_heat_out)
            sp_top_temp.append(node_i.top_temp)
            sp_thermocline.append(node_i.thermocline)
            sp_stored_energy.append(node_i.energy)
            node_i = node_i.next_node
        sp_soc = [(x-self.bottom_node.energy) / (self.top_node.energy-self.bottom_node.energy) * 100 
                    for x in sp_stored_energy]
        sp_time = list(range(self.params.horizon+1))
        
        # Plot the shortest path
        fig, ax = plt.subplots(2,1, sharex=True, figsize=(10,6))
        start = self.params.start_time.format('YYYY-MM-DD HH:mm')
        end = self.params.start_time.add(hours=self.params.horizon).format('YYYY-MM-DD HH:mm')
        fig.suptitle(f'From {start} to {end}\nCost: {round(self.initial_node.pathcost,2)} $', fontsize=10)
        
        # Top plot
        ax[0].step(sp_time, sp_hp_heat_out, where='post', color='tab:blue', alpha=0.6, label='HP')
        ax[0].step(sp_time, self.forecasts.load, where='post', color='tab:red', alpha=0.6, label='Load')
        ax[0].legend(loc='upper left')
        ax[0].set_ylabel('Heat [kWh]')
        ax[0].set_ylim([-0.5, 1.5*max(sp_hp_heat_out)])
        ax2 = ax[0].twinx()
        ax2.step(sp_time, self.forecasts.elec_price, where='post', color='gray', alpha=0.6, label='Elec price')
        ax2.legend(loc='upper right')
        ax2.set_ylabel('Electricity price [cts/kWh]')
        m = 0 if min(self.forecasts.elec_price)>0 else min(self.forecasts.elec_price)-5
        ax2.set_ylim([m,max(self.forecasts.elec_price)*1.3])
        
        # Bottom plot
        norm = Normalize(vmin=self.params.available_top_temps[0], vmax=self.params.available_top_temps[-1])
        cmap = matplotlib.colormaps['Reds']
        tank_top_colors = [cmap(norm(x)) for x in sp_top_temp]
        tank_bottom_colors = [cmap(norm(x-self.params.delta_T(x))) for x in sp_top_temp]
        sp_thermocline_reversed = [self.params.num_layers-x+1 for x in sp_thermocline]
        ax[1].bar(sp_time, sp_thermocline, bottom=sp_thermocline_reversed, color=tank_top_colors, alpha=0.7)
        ax[1].bar(sp_time, sp_thermocline_reversed, color=tank_bottom_colors, alpha=0.7)
        ax[1].set_xlabel('Time [hours]')
        ax[1].set_ylabel('Storage state')
        ax[1].set_ylim([0, self.params.num_layers])
        ax[1].set_yticks([])
        if len(sp_time)>10 and len(sp_time)<50:
            ax[1].set_xticks(list(range(0,len(sp_time)+1,2)))
        ax3 = ax[1].twinx()
        ax3.plot(sp_time, sp_soc, color='black', alpha=0.4, label='SoC')
        ax3.set_ylabel('State of charge [%]')
        ax3.set_ylim([-1,101])

        # Color bar
        boundaries = self.params.available_top_temps
        colors = [plt.cm.Reds(i/(len(boundaries)-1)) for i in range(len(boundaries))]
        cmap = ListedColormap(colors)
        norm = BoundaryNorm(boundaries, cmap.N, clip=True)
        sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
        cbar = plt.colorbar(sm, ax=ax, orientation='horizontal', fraction=0.06, pad=0.15, alpha=0.7)
        cbar.set_ticks(self.params.available_top_temps)
        cbar.set_label('Temperature [F]')
        
        plt.savefig('plot.png', dpi=130)
        plt.show()

    def export_excel(self):
        # Along the shortest path
        electricitiy_used, heat_delivered = [], []
        node_i = self.initial_node
        while node_i.next_node is not None:
            heat_to_store = node_i.next_node.energy - node_i.energy
            losses = self.params.storage_losses_percent/100*(node_i.energy-self.bottom_node.energy)
            if losses<self.energy_between_consecutive_states and losses>0 and self.forecasts.load[node_i.time_slice]==0:
                losses = self.energy_between_consecutive_states + 1/1e9
            heat_output_HP = heat_to_store + self.forecasts.load[node_i.time_slice] + losses
            cop = self.params.COP(oat=self.forecasts.oat[node_i.time_slice], lwt=node_i.next_node.top_temp)
            electricitiy_used.append(heat_output_HP / cop)
            heat_delivered.append(heat_output_HP)
            node_i = node_i.next_node
        
        # First dataframe: the Dijkstra graph
        dijkstra_pathcosts = {}
        dijkstra_pathcosts['Top Temp [F]'] = [x.top_temp for x in self.nodes_by_energy]
        dijkstra_pathcosts['Thermocline'] = [x.thermocline for x in self.nodes_by_energy]
        dijkstra_pathcosts['Index'] = list(range(1,len(self.nodes_by_energy)+1))
        dijkstra_nextnodes = dijkstra_pathcosts.copy()
        for h in range(self.params.horizon):
            dijkstra_pathcosts[h] = [round(x.pathcost,2) for x in sorted(self.nodes[h], key=lambda x: x.index)]
            dijkstra_nextnodes[h] = [x.next_node.index for x in sorted(self.nodes[h], key=lambda x: x.index)]
        dijkstra_pathcosts[self.params.horizon] = [0 for x in self.nodes[self.params.horizon]]
        dijkstra_nextnodes[self.params.horizon] = [np.nan for x in self.nodes[self.params.horizon]]
        dijkstra_pathcosts_df = pd.DataFrame(dijkstra_pathcosts)
        dijkstra_nextnodes_df = pd.DataFrame(dijkstra_nextnodes)
        
        # Second dataframe: the forecasts
        forecast_df = pd.DataFrame({'Forecast':['0'], 'Unit':['0'], **{h: [0.0] for h in range(self.params.horizon+1)}})
        forecast_df.loc[0] = ['Price - total'] + ['cts/kWh'] + self.forecasts.elec_price
        forecast_df.loc[1] = ['Price - distribution'] + ['cts/kWh'] + self.forecasts.dp
        forecast_df.loc[2] = ['Price - LMP'] + ['cts/kWh'] + self.forecasts.lmp
        forecast_df.loc[3] = ['Heating load'] + ['kW'] + [round(x,2) for x in self.forecasts.load]
        forecast_df.loc[4] = ['OAT'] + ['F'] + [round(x,2) for x in self.forecasts.oat]
        forecast_df.loc[5] = ['Required SWT'] + ['F'] + [round(x) for x in self.forecasts.rswt]
        
        # Third dataframe: the shortest path
        shortestpath_df = pd.DataFrame({'Shortest path':['0'], 'Unit':['0'], **{h: [0.0] for h in range(self.params.horizon+1)}})
        shortestpath_df.loc[0] = ['Electricity used'] + ['kWh'] + [round(x,3) for x in electricitiy_used] + [0]
        shortestpath_df.loc[1] = ['Heat delivered'] + ['kWh'] + [round(x,3) for x in heat_delivered] + [0]
        shortestpath_df.loc[2] = ['Cost - total'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.forecasts.elec_price)] + [0]
        shortestpath_df.loc[3] = ['Cost - distribution'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.forecasts.dp)] + [0]
        shortestpath_df.loc[4] = ['Cost - LMP'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.forecasts.lmp)] + [0]
        
        # Fourth dataframe: the results
        total_usd = round(self.initial_node.pathcost,2)
        total_elec = round(sum(electricitiy_used),2)
        total_heat = round(sum(heat_delivered),2)
        next_index = self.initial_node.next_node.index
        results = ['Cost ($)', total_usd, 'Electricity (kWh)', total_elec, 'Heat (kWh)', total_heat, 'Next step index', next_index]
        results_df = pd.DataFrame({'RESULTS':results})
        
        # Highlight shortest path
        highlight_positions = []
        node_i = self.initial_node
        while node_i.next_node is not None:
            highlight_positions.append((node_i.index+len(forecast_df)+len(shortestpath_df)+2, 3+node_i.time_slice))
            node_i = node_i.next_node
        highlight_positions.append((node_i.index+len(forecast_df)+len(shortestpath_df)+2, 3+node_i.time_slice))
        
        # Read the configuration file
        parameters = {}
        with open('parameters.conf', 'r') as file:
            for line in file:
                stripped_line = line.strip()
                if stripped_line and not stripped_line.startswith('#'):
                    key_value = stripped_line.split('=')
                    if len(key_value) == 2:
                        key = key_value[0].strip()
                        value = key_value[1].strip()
                        parameters[key] = value
        parameters_df = pd.DataFrame(list(parameters.items()), columns=['Variable', 'Value'])

        # Write to Excel
        os.makedirs('results', exist_ok=True)
        file_path = os.path.join('results', f'result_{self.params.now_for_file}.xlsx')
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            results_df.to_excel(writer, index=False, sheet_name='Pathcost')
            results_df.to_excel(writer, index=False, sheet_name='Next node')
            forecast_df.to_excel(writer, index=False, startcol=1, sheet_name='Pathcost')
            forecast_df.to_excel(writer, index=False, startcol=1, sheet_name='Next node')
            shortestpath_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+1, sheet_name='Pathcost')
            shortestpath_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+1, sheet_name='Next node')
            dijkstra_pathcosts_df.to_excel(writer, index=False, startrow=len(forecast_df)+len(shortestpath_df)+2, sheet_name='Pathcost')
            dijkstra_nextnodes_df.to_excel(writer, index=False, startrow=len(forecast_df)+len(shortestpath_df)+2, sheet_name='Next node')
            parameters_df.to_excel(writer, index=False, sheet_name='Parameters')
            pathcost_sheet = writer.sheets['Pathcost']
            nextnode_sheet = writer.sheets['Next node']
            parameters_sheet = writer.sheets['Parameters']
            plot_sheet = writer.book.create_sheet(title='Plot')
            plot_sheet.add_image(Image('plot.png'), 'A1')
            for row in pathcost_sheet['A1:A10']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
            for row in nextnode_sheet['A1:A10']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
            pathcost_sheet.column_dimensions['A'].width = 15
            pathcost_sheet.column_dimensions['B'].width = 15
            pathcost_sheet.column_dimensions['C'].width = 15
            nextnode_sheet.column_dimensions['A'].width = 15
            nextnode_sheet.column_dimensions['B'].width = 15
            nextnode_sheet.column_dimensions['C'].width = 15
            parameters_sheet.column_dimensions['A'].width = 40
            parameters_sheet.column_dimensions['B'].width = 70
            pathcost_sheet.freeze_panes = 'D14'
            nextnode_sheet.freeze_panes = 'D14'
            highlight_fill = PatternFill(start_color='72ba93', end_color='72ba93', fill_type='solid')
            for row in range(len(forecast_df)+len(shortestpath_df)+2):
                pathcost_sheet.cell(row=row+1, column=1).fill = highlight_fill
                nextnode_sheet.cell(row=row+1, column=1).fill = highlight_fill
            for row, col in highlight_positions:
                pathcost_sheet.cell(row=row+1, column=col+1).fill = highlight_fill
                nextnode_sheet.cell(row=row+1, column=col+1).fill = highlight_fill
        os.remove('plot.png')

g = DGraph()